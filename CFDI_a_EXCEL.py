from decimal import ROUND_HALF_UP as UP
from decimal import Decimal
from os import walk
from datetime import datetime
from tkinter.constants import W
import xml.etree.ElementTree as ET
from black import traceback
from joblib import PrintTime
from kivy import kivy_configure
from numpy import append
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from cfdiclient import Validacion
from tkinter import *
from tkinter import filedialog, ttk, messagebox

from sympy import apart

def GetFacturas(path_egresos_o_ingresos): #Regresa una lista de Facturas, bien sean ingresos o egresos
    list = []
    for root, dirs, xmls in walk(path_egresos_o_ingresos):
        for name in xmls:
            list.append(name)

    return list

def Ordenar_Facturas(path,facturas): #Regresa una lista de Facturas ya ordenadas por fecha, bien sean ingresos o egresos
    Lista_Facturas_con_fechas = []

    #Se relaciona los nombres de las facturas con su fecha
    for fact in facturas:
        tree = ET.parse(path + '/' + fact)
        root = tree.getroot()
        fecha = root.get('Fecha')
        Lista_Facturas_con_fechas.append({'Factura':fact, 'Fecha':fecha})

    #Se Ordena dichas facturas por fecha
    Lista_Facturas_Ordenadas = sorted(
        Lista_Facturas_con_fechas, key=lambda diccionario_fact: datetime.strptime(diccionario_fact['Fecha'], '%Y-%m-%dT%H:%M:%S'), reverse=False
    )

    #Limpia la lista donde se almacenan las facturas (Solo para no hacer otra lista)
    Lista_Facturas_con_fechas.clear()

    #Se almacenan las facturas ordenadas pero sin las fechas
    for fact in Lista_Facturas_Ordenadas:
        Lista_Facturas_con_fechas.append(fact['Factura']) 

    return Lista_Facturas_con_fechas

def Formar_Contenido(path,Facturas_ordenadas,Ingresos_o_Egresos):
    Contenido = []
    for factura in Facturas_ordenadas:
        
        #Crea una lista y la llena con NO Deducibles alojados en un archivo .txt
        NO_DEDUCIBLES = []
        archivo = open('NO_DEDUCIBLES.txt','r')
        for linea in archivo:
            NO_DEDUCIBLES.append(linea)

        tree = ET.parse(path + '/' + factura)
        root = tree.getroot()
        
        #--------------------DATOS GENERALES--------------------#
        Temp_Folios = []
        Folio_Fiscal = factura[:-4].upper()
        Temp_Folios.append(Folio_Fiscal)
        RFC_Emisor = root.find("{http://www.sat.gob.mx/cfd/3}Emisor").attrib['Rfc']
        RFC_Receptor = root.find("{http://www.sat.gob.mx/cfd/3}Receptor").attrib['Rfc']

        try:
            Razon_Social_Emisor = root.find("{http://www.sat.gob.mx/cfd/3}Emisor").attrib['Nombre']
            Razon_Social_Receptor = root.find("{http://www.sat.gob.mx/cfd/3}Receptor").attrib['Nombre']
        except KeyError:
            Razon_Social_Receptor = 'CONSULTAR EN LA PÁGINA'
            Razon_Social_Emisor = 'CONSULTAR EN LA PÁGINA'

        Fecha_Emision = root.get('Fecha')
        FechaTimbrado = root.find("{http://www.sat.gob.mx/cfd/3}Complemento").find('{http://www.sat.gob.mx/TimbreFiscalDigital}TimbreFiscalDigital').attrib['FechaTimbrado']
        RfcProvCertif = root.find("{http://www.sat.gob.mx/cfd/3}Complemento").find('{http://www.sat.gob.mx/TimbreFiscalDigital}TimbreFiscalDigital').attrib['RfcProvCertif']
        Total = float(root.get('Total'))
        SubTotal = float(root.get('SubTotal'))
        Folio = root.get('Folio')
        Descuento = root.get('Descuento')
        if(Descuento == None): Descuento = 0
        MetodoPago = root.get('MetodoPago')
        Efecto_Comprobante = root.find("{http://www.sat.gob.mx/cfd/3}Conceptos").find('{http://www.sat.gob.mx/cfd/3}Concepto').attrib['Descripcion']
        if (Efecto_Comprobante != 'Pago'): Efecto_Comprobante = 'Ingreso' 
        estado = Validacion().obtener_estado(RFC_Emisor, RFC_Receptor, str(Total), Folio_Fiscal)
        
        try:
            Estatus_Cancelacion = estado['es_cancelable']
            Estado_Comprobante = estado['estado']

            if(Estatus_Cancelacion and Estado_Comprobante == None):
                Estatus_Cancelacion = "REVISAR EN LA PAGINA"
                Estado_Comprobante = "REVISAR EN LA PAGINA"

        except:
            messagebox.showerror(title='Error', message='No hay respuesta del servidor')
        ############################################################ 
        Tasa_0 = False
        Tasa_16 = False
        try:
            child_impuesto = root.find("{http://www.sat.gob.mx/cfd/3}Impuestos")
            Impuesto = float(child_impuesto.attrib['TotalImpuestosTrasladados'])
            Exento = False

            #REVISA Los tipos de impuestos, si hay al 16 y/o al 0%
            traslados = root.find("{http://www.sat.gob.mx/cfd/3}Impuestos").find('{http://www.sat.gob.mx/cfd/3}Traslados')
            for i in traslados:
                Tipo_tasa = list(i.attrib.values())[2]
                if(Tipo_tasa == "0.160000"):
                    Tasa_16 = True
                if (Tipo_tasa == "0.000000"):
                    Tasa_0 = True
                else:
                    if(Impuesto == 0):
                        Tasa_0 = True
                    else:
                        Tasa_16 = True

        except(AttributeError, KeyError) as Error:
            Impuesto = 0.0
            Exento = True
        #-------------------DATOS SOBRE IMPUESTOS--------------------# 
        #Solo son de prueba:
        #Estatus_Cancelacion = 'Prueba' 
        #Estado_Comprobante = 'Vigente' 
        
        if (Ingresos_o_Egresos):
            if(Estado_Comprobante == 'Vigente'):
                Fecha_Cuadro = Fecha_Emision[:-9]
                if(Razon_Social_Receptor == 'PUBLICO EN GENERAL'):
                    SubTotal_Cuadro, Al_16, Al_0 = [0.0,0.0,SubTotal]
                else:
                    SubTotal_Cuadro, Al_16, Al_0 = [SubTotal,Impuesto,0.0]
            else:
                Fecha_Cuadro, Folio, Razon_Social_Receptor, SubTotal_Cuadro, Al_16, Al_0 = ["","","",0.0,0.0,0.0]
            
            Contenido.append([Folio_Fiscal, RFC_Emisor,Razon_Social_Emisor,RFC_Receptor,Razon_Social_Receptor,Fecha_Emision,FechaTimbrado,RfcProvCertif,Total,Efecto_Comprobante,Estatus_Cancelacion, Estado_Comprobante, '',Fecha_Cuadro,Folio,Razon_Social_Receptor,SubTotal_Cuadro,Al_16,Al_0])
        else:  
            def Calcular_Impuesto(Dis, T_16, T_0,Total):
                #Valores predeterminados
                SubTotal_Resultado = 0
                Al_16 = 0
                Al_0 = 0
                
                # Operaciones para Tasas al 16%
                if(T_16 and T_0 == False):
                    if(Dis):# Verificar si tiene descuento
                        #Operaciones cuando tiene descuento.
                        SubTotal_Calculado = Total/1.16
                        Impuesto_Calculado = SubTotal_Calculado * 0.16 

                        #Asignar resultados
                        SubTotal_Resultado = SubTotal_Calculado
                        Al_16 = Impuesto_Calculado
                    else:
                        #Operaciones cuando NO tiene descuento
                        SubTotal_Resultado = SubTotal #checar var
                        Al_16 = Impuesto #checar var

                # Operaciones para Tasas al 0.0%
                elif (T_16 == False and T_0):
                    if(Dis): #Verificar si tiene descuento
                        #Operaciones cuando tiene descuento
                        Al_0 = Total #Poner el total en la columna de Al 0%
                    else:
                        Al_0 = SubTotal #Poner el subtotal en la columna de Al 0%
                # Operaciones para Tasas al 16% y al 0.0%
                elif(T_16 and T_0):
                    #Calculos previos
                    SubTotal_Calculado = Total/1.16
                    Impuesto_Calculado = SubTotal_Calculado * 0.16

                    #Calcular Subtotal al 16%
                    SubTotal_16 = Impuesto * SubTotal_Calculado / Impuesto_Calculado

                    #Calcular Subtotal al 0%
                    SubTotal_0 = Total - SubTotal_16 - Impuesto

                    #Asignar Datos
                    SubTotal_Resultado = SubTotal_16
                    Al_16 = Impuesto
                    Al_0 = SubTotal_0

                SubTotal_Resultado = float(Decimal(str(SubTotal_Resultado)).quantize(Decimal("0.00"), rounding=UP))
                Al_16 = float(Decimal(str(Al_16)).quantize(Decimal("0.00"), rounding=UP))
                Al_0 = float(Decimal(str(Al_0)).quantize(Decimal("0.00"), rounding=UP))

                return (SubTotal_Resultado, Al_16, Al_0)
        
            if(MetodoPago != 'PPD'):
                if not(Razon_Social_Emisor in NO_DEDUCIBLES):
                    Fecha_Cuadro = Fecha_Emision[:-9]
                    if(Exento):
                        SubTotal_Cuadro, Al_16, Al_0, Exento_value, Nota = [0.0,0.0,0.0,SubTotal,'EXENTO']
                    else:
                        if(float(Descuento) == 0):
                            SubTotal_Cuadro,Al_16, Al_0 = Calcular_Impuesto(False, Tasa_16, Tasa_0, Total)
                            Nota = ''
                            Exento_value = 0.0
                        else:
                            SubTotal_Cuadro,Al_16, Al_0 = Calcular_Impuesto(True, Tasa_16, Tasa_0, Total)
                            Nota = ''
                            Exento_value = 0.0
                else:
                    Fecha_Cuadro, SubTotal_Cuadro, Al_16, Al_0, Exento_value, Nota = ['',0.0,0.0,0.0,0.0,'NO DEDUCIBLE']

            else:
                Fecha_Cuadro, SubTotal_Cuadro, Al_16, Al_0, Exento_value, Nota = ['',0.0,0.0,0.0,0.0,'PPD']

            Contenido.append([Folio_Fiscal, RFC_Emisor,Razon_Social_Emisor,RFC_Receptor,Razon_Social_Receptor,Fecha_Emision,FechaTimbrado,RfcProvCertif,Total,Efecto_Comprobante,Estatus_Cancelacion, Estado_Comprobante, '',Fecha_Cuadro,SubTotal_Cuadro,Al_16,Al_0,Exento_value,'',Nota]) 
        Barra_Progreso['value'] += 1
        FrameProgreso.update()
    return Contenido

def Insertar_Contenido(Contenido, Excel_name, Sheet_name, Ingresos_o_Egresos):
    validador = True

    #Crear una nueva hoja con el nombre elegido
    wb = load_workbook(Excel_name)
    ws = wb.create_sheet(Sheet_name)

    #Definir los encabezados
    Encabezados_Egresos = ['Folio Fiscal', 'RFC Emisor', 'Nombre o Razón Social del Emisor', 'RFC del Receptor', 'Nombre o Razón Social del Receptor', 'Fecha de Emisión', 'Fecha de Certificación', 'PAC que Certifición', 'Total','Efecto del Comprobante','Estatus de cancelación','Estado del Comprobante','','FECHA','SUBTOTAL','AL 16%','AL 0%','EXENTOS','TOTAL','NOTAS']
    Encabezados_Ingresos = Encabezados_Egresos[:14] + ["FOLIO", "CLIENTE", 'SUBTOTAL', "AL 16%", "AL 0%", "TOTAL"]
    
    #Definir estilos de bordes
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    button_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='double'))

    #Ingresar datos según sea ingreso o egreso
        # 1 Ingreso
        # 0 Egreso
    if(Ingresos_o_Egresos):
        ws.append(Encabezados_Ingresos) #Agrear los encabezados
        for line in Contenido:
            ws.append(line)
        Dim_cuadro = ws.max_row #Definir la dimensión maxima de las celdas
    
        #Agregar operaciones de suma a la columna de total
        for row in range(2,Dim_cuadro+1):
            ws[f'T{row}'] = f'=SUM(Q{row}:S{row})'
        #Agregar operaciones de suma al final del cuadro
        ws[f'Q{Dim_cuadro+1}'] = f'=SUM(Q2:Q{Dim_cuadro})'
        ws[f'R{Dim_cuadro+1}'] = f'=SUM(R2:R{Dim_cuadro})'
        ws[f'S{Dim_cuadro+1}'] = f'=SUM(S2:S{Dim_cuadro})'
        ws[f'T{Dim_cuadro+1}'] = f'=SUM(T2:T{Dim_cuadro})'
        
    ###########################################################################
    else:
        
        ws.append(Encabezados_Egresos) #Agrear los encabezados
    
        #Insertar el contenido
        for line in Contenido:
            ws.append(line)

        Dim_cuadro = ws.max_row #Definir la dimensión maxima de las celdas
        #Agregar operaciones de suma a la columna de total
        for row in range(2,Dim_cuadro+1):
            ws[f'S{row}'] = f'=SUM(O{row}:R{row})'

        #Agregar operaciones de suma al final del cuadro
        ws[f'O{Dim_cuadro+1}'] = f'=SUM(O2:O{Dim_cuadro})'
        ws[f'P{Dim_cuadro+1}'] = f'=SUM(P2:P{Dim_cuadro})'
        ws[f'Q{Dim_cuadro+1}'] = f'=SUM(Q2:Q{Dim_cuadro})'
        ws[f'R{Dim_cuadro+1}'] = f'=SUM(R2:R{Dim_cuadro})'
        ws[f'S{Dim_cuadro+1}'] = f'=SUM(S2:S{Dim_cuadro})'
    
    ########################################################################

    #Agregar los bordes para resaltar el resumen de las facturas
    for lin in range(1,Dim_cuadro+1):
        for number in range(0,7):
            ws.cell(row=lin, column=14+number).border = thin_border
            if(lin == Dim_cuadro): ws.cell(row=lin, column=14+number).border = button_border
    

    #Hacer negritas el encabezado
    bold_font = Font(size=12, bold=True)
    for cell in ws[1:1]:
        cell.font = bold_font


   ### RETO A IMPLEMENTAR: SEPARAR POR PAGINAS Y POR MES
    # with_spaces = len(Contenido) / 15 + len(Contenido) 
    # with_spaces = int(round(with_spaces,0)) 
    # C0 = 0
    # for line in range(with_spaces):
    #     if(line%16 == 0):
    #         ws.append([])
    #         pass
    #     else:
    #         ws.append(Contenido[C0])
    #         C0 += 1
    #ws.move_range(f"A{ws.max_row}:T{ws.max_row}", rows=-ws.max_row+1, cols=0)
    
    #Hacer los resultados de las sumas de cada columna en negritas
    for cell in ws[Dim_cuadro+1:Dim_cuadro+1]:
        cell.font = Font(bold=True)

    #Dejar estáticos los encabezados
    ws.freeze_panes = ws['A2']
    wb.save(Excel_name)

    return validador

def Ejecutar(Nombre_excel, Nombre_Sheet, path, Ingreso_o_Egreso):
    try:
        Lista_Facturas = GetFacturas(path)        
        Facturas_ordenadas = Ordenar_Facturas(path, Lista_Facturas)
        Barra_Progreso["maximum"] = len(Lista_Facturas)
        Contenido = Formar_Contenido(path, Facturas_ordenadas,Ingreso_o_Egreso)
        Insertar_Contenido(Contenido, Nombre_excel, Nombre_Sheet,Ingreso_o_Egreso)
        mensaje = f'PROCESO TERMINADO.\n {len(Lista_Facturas)}  Facturas Procesadas Exitosamente'
        messagebox.showinfo(message=mensaje, title='ESTADO')
        Barra_Progreso['value'] = 0
        Barra_Progreso.update()
        BotProcesar['state']  = 'normal'
    
    except Exception as cause:
        messagebox.showerror(message='Algo Salió Mal :(\n-Revise operacion ( Emitidas o Recibidas, es correcto ? \n-Si los datos son correctos', title='ERROR')
        Barra_Progreso['value'] = 0
        Barra_Progreso.update()
        BotProcesar['state']  = 'normal'

        traceback.print_exc()

def Procesar_Entradas():
    Tipo_Calculo = Op.get() #Selector
    excel_name = EntExcelpath.get()
    path_XML = EntXMLpath.get()
    Hoja_nombre = Entws.get()
    BotProcesar['state'] = 'disabled'
    if(Hoja_nombre == ''):
        messagebox.showerror(message='Algo Salió Mal :(\n-Revise operacion ( Emitidas o Recibidas, es correcto ? ) \n-Si los datos son correctos', title='ERROR')
    else:
        if (Tipo_Calculo == 0):
            Tipo_Calculo = 1
            Ejecutar(excel_name,Hoja_nombre ,path_XML,Tipo_Calculo)
        else:
            Tipo_Calculo = 0 
            Ejecutar(excel_name,Hoja_nombre ,path_XML,Tipo_Calculo)

def Clear_Entradas():
    EntExcelpath.delete(0, END)
    EntXMLpath.delete(0, END)
    Entws.delete(0, END)

def Buscar_path_XML():
    R2.set(filedialog.askdirectory())

def Buscar_path_Excel():
    Ruta.set(filedialog.askopenfilename(title="Ubicación del Excel", filetypes=[("Excel (.xlsx)","*.xlsx")]))


Window = Tk()
Window.title('CFDI a EXCEL v1.1 Alpha')
Window.resizable(0,0)
Window.iconbitmap('Calcu.ico')

#Formar contenedor
FrameGeneral = Frame(Window)
FrameGeneral.grid(row=0, column=0)

#Frames:
FrameProgreso = Frame(FrameGeneral)
FrameProgreso.grid(row=0, column=0)

FrameOpciones = Frame(FrameGeneral)
FrameOpciones.grid(row=1, column=0)#Frame para selccionar Egreso o Ingreso

FrameEgreso = Frame(FrameGeneral) #Mal nombrado. Es el Frame contenido
FrameEgreso.grid(row=2, column=0)

FrameEjecutar = Frame(FrameGeneral)
FrameEjecutar.grid(row=3, column=0, pady=5)

#Frame Opciones:
Op = IntVar()
Op.set(None)

OpcionIngreso = Radiobutton(FrameOpciones,text='Emitidas', variable= Op, value=0)
OpcionIngreso.grid(row=0, column=0)

OpcionEgreso = Radiobutton(FrameOpciones, text='Recibidas', variable=Op,value=1)
OpcionEgreso.grid(row=0, column=1)

#Frame Egreso:
lblExcelpath = Label(FrameEgreso,text='Ubicación Excel:')
lblExcelpath.grid(row=0, column=0)

Ruta = StringVar()
Ruta.set('')
EntExcelpath = Entry(FrameEgreso)
EntExcelpath.config(width=60, justify='center', textvariable=Ruta)
EntExcelpath.grid(row=0, column=1)

BotExcelpath = Button(FrameEgreso, text='...', command=Buscar_path_Excel)
BotExcelpath.grid(row=0, column=2)

#------------------------------------------------------------------------------#
lblXMLpath = Label(FrameEgreso, text="Ubicación CFDI's:")
lblXMLpath.grid(row= 1, column=0)

R2 = StringVar()
R2.set('')
EntXMLpath = Entry(FrameEgreso)
EntXMLpath.config(width=60, justify='center', textvariable=R2)
EntXMLpath.grid(row=1, column=1)

BotXMLpath = Button(FrameEgreso, text='...',command=Buscar_path_XML)
BotXMLpath.grid(row=1, column=2)

#------------------------------------------------------------------------------#
lblws = Label(FrameEgreso, text='Nombre Hoja')
lblws.grid(row=2, column=0)

ws_name = StringVar()
ws_name.set('')
Entws = Entry(FrameEgreso, textvariable=ws_name)
Entws.config(justify='center',width=30 )
Entws.grid(row=2, column=1, sticky=W)

#------------------------------------------------------------------------------#

#Frame Egreso:
BotProcesar = Button(FrameEjecutar, text='Procesar',command=Procesar_Entradas)
BotProcesar.config(fg='green')
BotProcesar.grid(row=0, column=0, padx=1)

Botclear = Button(FrameEjecutar, text='Limpiar',command=Clear_Entradas)
Botclear.config(fg='red')
Botclear.grid(row=0, column=1, padx=1)

#Frame Progreso:
Barra_Progreso = ttk.Progressbar(FrameProgreso, length=490, mode='determinate', maximum=100)
Barra_Progreso.grid(row=0, column=0)

Window.mainloop()